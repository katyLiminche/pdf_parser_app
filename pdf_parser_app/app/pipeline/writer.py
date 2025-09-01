"""
Excel writer module for exporting parsed data
"""

import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

logger = logging.getLogger(__name__)

class ExcelWriter:
    """Excel export functionality"""
    
    def __init__(self, backup_folder: str = "backups"):
        self.backup_folder = Path(backup_folder)
        self.backup_folder.mkdir(parents=True, exist_ok=True)
    
    def write_to_template(self, 
                         template_path: str, 
                         output_path: str, 
                         items: List[Dict[str, Any]], 
                         sheet_name: str = 'Raw_imports',
                         create_backup: bool = True) -> Optional[str]:
        """
        Write items to Excel template
        
        Args:
            template_path: Path to Excel template
            output_path: Path for output file
            items: List of items to export
            sheet_name: Sheet name to write to
            create_backup: Whether to create backup of template
            
        Returns:
            Path to backup file if created, None otherwise
        """
        try:
            template_path = Path(template_path)
            output_path = Path(output_path)
            
            # Create backup if requested
            backup_path = None
            if create_backup and template_path.exists():
                backup_path = self._create_backup(template_path)
            
            # Load workbook
            if template_path.exists():
                wb = load_workbook(template_path)
            else:
                # Create new workbook if template doesn't exist
                wb = load_workbook()
                logger.warning(f"Template {template_path} not found, creating new workbook")
            
            # Ensure sheet exists
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
                logger.info(f"Created new sheet: {sheet_name}")
            
            ws = wb[sheet_name]
            
            # Write data
            self._write_items_to_sheet(ws, items)
            
            # Save output file
            wb.save(output_path)
            logger.info(f"Exported {len(items)} items to {output_path}")
            
            return str(backup_path) if backup_path else None
            
        except Exception as e:
            logger.error(f"Failed to write to Excel: {e}")
            raise
    
    def _create_backup(self, template_path: Path) -> Path:
        """Create backup of template file"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f"{template_path.stem}_backup_{timestamp}{template_path.suffix}"
        backup_path = self.backup_folder / backup_name
        
        try:
            shutil.copy2(template_path, backup_path)
            logger.info(f"Created backup: {backup_path}")
            return backup_path
        except Exception as e:
            logger.error(f"Failed to create backup: {e}")
            raise
    
    def _write_items_to_sheet(self, worksheet, items: List[Dict[str, Any]]):
        """Write items to worksheet"""
        if not items:
            logger.warning("No items to write")
            return
        
        # Define column mapping
        column_mapping = {
            'supplier': 'A',
            'name': 'B', 
            'qty': 'C',
            'unit': 'D',
            'price': 'E',
            'currency': 'F',
            'total': 'G',
            'sku': 'H',
            'source_file': 'I',
            'confidence': 'J'
        }
        
        # Find starting row (after existing data)
        start_row = worksheet.max_row + 1
        
        # Write header if sheet is empty
        if start_row == 1:
            headers = list(column_mapping.keys())
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                cell.font = cell.font.copy(bold=True)
            start_row = 2
        
        # Write items
        for row_idx, item in enumerate(items, start_row):
            try:
                # Map item data to columns
                row_data = {
                    'supplier': item.get('supplier', ''),
                    'name': item.get('name', ''),
                    'qty': item.get('qty', ''),
                    'unit': item.get('unit', ''),
                    'price': item.get('price', ''),
                    'currency': item.get('currency', ''),
                    'total': item.get('total', ''),
                    'sku': item.get('sku', item.get('sku_suggestion', '')),
                    'source_file': item.get('source_file', ''),
                    'confidence': item.get('confidence_score', '')
                }
                
                # Write to worksheet
                for col_idx, (field, value) in enumerate(row_data.items(), 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Format numeric cells
                    if field in ['qty', 'price', 'total'] and value is not None:
                        try:
                            cell.number_format = '#,##0.00'
                        except:
                            pass
                    
                    # Color code confidence
                    if field == 'confidence' and value is not None:
                        try:
                            confidence = float(value)
                            if confidence >= 0.9:
                                cell.fill = self._get_fill_color('00FF00')  # Green
                            elif confidence >= 0.7:
                                cell.fill = self._get_fill_color('FFFF00')  # Yellow
                            else:
                                cell.fill = self._get_fill_color('FF0000')  # Red
                        except:
                            pass
                
            except Exception as e:
                logger.warning(f"Failed to write row {row_idx}: {e}")
                continue
        
        # Auto-adjust column widths
        self._auto_adjust_columns(worksheet, column_mapping)
        
        logger.info(f"Wrote {len(items)} items starting from row {start_row}")
    
    def _get_fill_color(self, color_code: str):
        """Get fill color for cells"""
        from openpyxl.styles import PatternFill
        return PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')
    
    def _auto_adjust_columns(self, worksheet, column_mapping: Dict[str, str]):
        """Auto-adjust column widths"""
        try:
            for col_idx in range(1, len(column_mapping) + 1):
                column_letter = worksheet.cell(row=1, column=col_idx).column_letter
                worksheet.column_dimensions[column_letter].auto_size = True
        except Exception as e:
            logger.debug(f"Column auto-adjust failed: {e}")
    
    def export_to_dataframe(self, items: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        Convert items to DataFrame for export
        
        Args:
            items: List of items to convert
            
        Returns:
            DataFrame with items
        """
        if not items:
            return pd.DataFrame()
        
        # Convert to DataFrame
        df = pd.DataFrame(items)
        
        # Reorder columns for Excel export
        column_order = [
            'supplier', 'name', 'qty', 'unit', 'price', 
            'currency', 'total', 'sku', 'source_file', 'confidence_score'
        ]
        
        # Add missing columns with default values
        for col in column_order:
            if col not in df.columns:
                df[col] = ''
        
        # Reorder and select columns
        df = df[column_order]
        
        # Rename confidence column
        df = df.rename(columns={'confidence_score': 'confidence'})
        
        return df
    
    def write_dataframe_to_excel(self, 
                                df: pd.DataFrame, 
                                output_path: str, 
                                sheet_name: str = 'Raw_imports') -> bool:
        """
        Write DataFrame directly to Excel
        
        Args:
            df: DataFrame to export
            output_path: Path for output file
            sheet_name: Sheet name
            
        Returns:
            True if successful
        """
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            logger.info(f"Exported DataFrame to {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to export DataFrame: {e}")
            return False
    
    def get_template_info(self, template_path: str) -> Dict[str, Any]:
        """
        Get information about Excel template
        
        Args:
            template_path: Path to template
            
        Returns:
            Template information
        """
        try:
            template_path = Path(template_path)
            if not template_path.exists():
                return {'error': 'Template file not found'}
            
            wb = load_workbook(template_path, read_only=True)
            
            info = {
                'file_size': template_path.stat().st_size,
                'sheets': wb.sheetnames,
                'has_raw_imports': 'Raw_imports' in wb.sheetnames
            }
            
            if 'Raw_imports' in wb.sheetnames:
                ws = wb['Raw_imports']
                info['raw_imports_rows'] = ws.max_row
                info['raw_imports_cols'] = ws.max_column
            
            wb.close()
            return info
            
        except Exception as e:
            logger.error(f"Failed to get template info: {e}")
            return {'error': str(e)}
