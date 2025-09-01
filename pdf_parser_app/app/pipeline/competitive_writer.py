"""
Специализированный Excel writer для конкурентной процедуры
"""

import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

logger = logging.getLogger(__name__)

class CompetitiveExcelWriter:
    """Excel writer для конкурентной процедуры закупок"""
    
    def __init__(self, backup_folder: str = "backups"):
        self.backup_folder = Path(backup_folder)
        self.backup_folder.mkdir(parents=True, exist_ok=True)
        
        # Стили для конкурентной процедуры
        self.header_style = {
            'font': Font(bold=True, color="FFFFFF"),
            'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
        self.data_style = {
            'font': Font(size=11),
            'alignment': Alignment(horizontal="left", vertical="center"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
        self.number_style = {
            'font': Font(size=11),
            'alignment': Alignment(horizontal="right", vertical="center"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def write_competitive_data(self, 
                              template_path: str, 
                              output_path: str, 
                              items: List[Dict[str, Any]], 
                              sheet_name: str = 'Raw_imports',
                              create_backup: bool = True) -> Optional[str]:
        """
        Запись данных конкурентной процедуры в Excel
        
        Args:
            template_path: Путь к шаблону Excel
            output_path: Путь для выходного файла
            items: Список позиций
            sheet_name: Имя листа
            create_backup: Создавать ли резервную копию
            
        Returns:
            Путь к резервной копии или None
        """
        try:
            template_path = Path(template_path)
            output_path = Path(output_path)
            
            # Создаем резервную копию
            backup_path = None
            if create_backup and template_path.exists():
                backup_path = self._create_backup(template_path)
            
            # Загружаем или создаем рабочую книгу
            if template_path.exists():
                wb = load_workbook(template_path)
            else:
                wb = load_workbook()
                logger.warning(f"Шаблон {template_path} не найден, создается новая книга")
            
            # Обеспечиваем наличие листа
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
                logger.info(f"Создан новый лист: {sheet_name}")
            
            ws = wb[sheet_name]
            
            # Записываем данные
            self._write_competitive_items(ws, items)
            
            # Сохраняем файл
            wb.save(output_path)
            logger.info(f"Экспортировано {len(items)} позиций в {output_path}")
            
            return str(backup_path) if backup_path else None
            
        except Exception as e:
            logger.error(f"Ошибка записи в Excel: {e}")
            raise
    
    def _create_backup(self, template_path: Path) -> Path:
        """Создание резервной копии шаблона"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f"{template_path.stem}_competitive_backup_{timestamp}{template_path.suffix}"
        backup_path = self.backup_folder / backup_name
        
        try:
            shutil.copy2(template_path, backup_path)
            logger.info(f"Создана резервная копия: {backup_path}")
            return backup_path
        except Exception as e:
            logger.error(f"Ошибка создания резервной копии: {e}")
            raise
    
    def _write_competitive_items(self, worksheet, items: List[Dict[str, Any]]):
        """Запись позиций конкурентной процедуры в лист"""
        if not items:
            logger.warning("Нет позиций для записи")
            return
        
        # Определяем колонки для конкурентной процедуры
        columns = [
            'Поставщик', 'Наименование', 'Количество', 'Единица', 
            'Цена за единицу', 'Валюта', 'Общая стоимость', 'SKU', 
            'Источник', 'Уверенность', 'Дата обработки'
        ]
        
        # Находим начальную строку
        start_row = worksheet.max_row + 1
        
        # Записываем заголовки если лист пустой
        if start_row == 1:
            for col_idx, header in enumerate(columns, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                self._apply_style(cell, self.header_style)
            start_row = 2
        
        # Записываем позиции
        for row_idx, item in enumerate(items, start_row):
            try:
                # Подготавливаем данные строки
                row_data = [
                    item.get('supplier', ''),           # Поставщик
                    item.get('name', ''),               # Наименование
                    item.get('qty', ''),                # Количество
                    item.get('unit', ''),               # Единица
                    item.get('price', ''),              # Цена за единицу
                    item.get('currency', ''),           # Валюта
                    item.get('total', ''),              # Общая стоимость
                    item.get('sku', ''),                # SKU
                    item.get('source', ''),             # Источник
                    item.get('confidence', ''),         # Уверенность
                    datetime.now().strftime('%Y-%m-%d') # Дата обработки
                ]
                
                # Записываем в лист
                for col_idx, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Применяем стили
                    if col_idx in [3, 5, 7]:  # Количество, Цена, Сумма
                        self._apply_style(cell, self.number_style)
                        if isinstance(value, (int, float)) and value != 0:
                            cell.number_format = '#,##0.00'
                    else:
                        self._apply_style(cell, self.data_style)
                    
                    # Цветовое кодирование уверенности
                    if col_idx == 10 and value is not None:  # Колонка уверенности
                        try:
                            confidence = float(value)
                            if confidence >= 0.9:
                                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                            elif confidence >= 0.7:
                                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        except:
                            pass
                
            except Exception as e:
                logger.warning(f"Ошибка записи строки {row_idx}: {e}")
                continue
        
        # Автоматически подгоняем ширину колонок
        self._auto_adjust_columns(worksheet, len(columns))
        
        logger.info(f"Записано {len(items)} позиций начиная со строки {start_row}")
    
    def _apply_style(self, cell, style_dict):
        """Применение стиля к ячейке"""
        try:
            for attr, value in style_dict.items():
                setattr(cell, attr, value)
        except Exception as e:
            logger.debug(f"Ошибка применения стиля: {e}")
    
    def _auto_adjust_columns(self, worksheet, num_columns):
        """Автоматическая подгонка ширины колонок"""
        try:
            for col_idx in range(1, num_columns + 1):
                column_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[column_letter].auto_size = True
        except Exception as e:
            logger.debug(f"Ошибка автоподгонки колонок: {e}")
    
    def export_to_dataframe(self, items: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        Конвертация позиций в DataFrame для экспорта
        
        Args:
            items: Список позиций
            
        Returns:
            DataFrame с данными
        """
        if not items:
            return pd.DataFrame()
        
        # Подготавливаем данные для DataFrame
        df_data = []
        for item in items:
            df_data.append({
                'Поставщик': item.get('supplier', ''),
                'Наименование': item.get('name', ''),
                'Количество': item.get('qty', ''),
                'Единица': item.get('unit', ''),
                'Цена за единицу': item.get('price', ''),
                'Валюта': item.get('currency', ''),
                'Общая стоимость': item.get('total', ''),
                'SKU': item.get('sku', ''),
                'Источник': item.get('source', ''),
                'Уверенность': item.get('confidence', ''),
                'Дата обработки': datetime.now().strftime('%Y-%m-%d')
            })
        
        return pd.DataFrame(df_data)
    
    def write_dataframe_to_excel(self, 
                                df: pd.DataFrame, 
                                output_path: str, 
                                sheet_name: str = 'Raw_imports') -> bool:
        """
        Запись DataFrame в Excel
        
        Args:
            df: DataFrame для экспорта
            output_path: Путь для выходного файла
            sheet_name: Имя листа
            
        Returns:
            True если успешно
        """
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            logger.info(f"DataFrame экспортирован в {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка экспорта DataFrame: {e}")
            return False
    
    def get_template_info(self, template_path: str) -> Dict[str, Any]:
        """
        Получение информации о шаблоне Excel
        
        Args:
            template_path: Путь к шаблону
            
        Returns:
            Информация о шаблоне
        """
        try:
            template_path = Path(template_path)
            if not template_path.exists():
                return {'error': 'Файл шаблона не найден'}
            
            wb = load_workbook(template_path, read_only=True)
            
            info = {
                'file_size': template_path.stat().st_size,
                'sheets': wb.sheetnames,
                'has_raw_imports': 'Raw_imports' in wb.sheetnames
            }
            
            if 'Raw_imports' in wb.sheetnames:
                ws = wb['Raw_imports']
                info['raw_imports_rows'] = ws.max_row
                info['raw_imports_cols'] = ws.max_column
            
            wb.close()
            return info
            
        except Exception as e:
            logger.error(f"Ошибка получения информации о шаблоне: {e}")
            return {'error': str(e)}
