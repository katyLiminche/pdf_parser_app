"""
Универсальный Excel writer для комбинированного подхода
"""

import os
import shutil
from datetime import datetime
from typing import List, Dict, Any, Optional
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class UniversalExcelWriter:
    """Универсальный Excel writer для комбинированного подхода"""
    
    def __init__(self, template_path: str = None):
        self.template_path = template_path
        self.sheet_name = 'Raw_imports'
        
        # Цвета для форматирования
        self.colors = {
            'header': '366092',  # Синий
            'supplier_profile': 'C6EFCE',  # Зеленый для профилей поставщиков
            'universal': 'FFEB9C',  # Желтый для универсального парсера
            'commercial': 'FFC7CE',  # Красный для коммерческого парсера
            'invoice': 'D9E1F2',  # Голубой для счетов
            'competitive': 'F2F2F2',  # Серый для конкурентного
            'border': '000000'  # Черный для границ
        }
    
    def write_items_to_excel(self, items: List[Dict[str, Any]], output_path: str, 
                            parser_info: Dict[str, Any] = None) -> str:
        """
        Записывает элементы в Excel файл с улучшенным форматированием
        
        Args:
            items: Список элементов для записи
            output_path: Путь к выходному файлу
            parser_info: Информация о парсере
            
        Returns:
            Путь к созданному файлу
        """
        try:
            # Создаем новый файл или используем шаблон
            if self.template_path and os.path.exists(self.template_path):
                # Копируем шаблон
                shutil.copy2(self.template_path, output_path)
                wb = load_workbook(output_path)
            else:
                # Создаем новый файл
                wb = load_workbook()
                wb.remove(wb.active)  # Удаляем дефолтный лист
            
            # Получаем или создаем лист
            if self.sheet_name in wb.sheetnames:
                ws = wb[self.sheet_name]
            else:
                ws = wb.create_sheet(self.sheet_name)
            
            # Очищаем старые данные (если есть)
            self._clear_sheet(ws)
            
            # Записываем заголовки
            self._write_headers(ws)
            
            # Записываем данные
            start_row = 2  # Начинаем с 2-й строки (после заголовков)
            for i, item in enumerate(items, start_row):
                self._write_item_row(ws, item, i, parser_info)
            
            # Применяем форматирование
            self._apply_formatting(ws, len(items))
            
            # Сохраняем файл
            wb.save(output_path)
            
            logger.info(f"Данные записаны в Excel файл: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Ошибка записи в Excel: {e}")
            raise
    
    def _clear_sheet(self, ws):
        """Очищает лист от старых данных"""
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
    
    def _write_headers(self, ws):
        """Записывает заголовки"""
        headers = [
            '№', 'Поставщик', 'Парсер', 'Наименование', 'Артикул', 
            'Количество', 'Единица', 'Цена', 'Валюта', 'Сумма',
            'Уверенность', 'Источник', 'Дата обработки'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['header'], 
                                   end_color=self.colors['header'], 
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _write_item_row(self, ws, item: Dict[str, Any], row: int, parser_info: Dict[str, Any] = None):
        """Записывает строку с элементом"""
        # Определяем цвет в зависимости от парсера
        parser_type = item.get('parser_used', 'unknown')
        fill_color = self._get_parser_color(parser_type)
        
        # Записываем данные
        data = [
            row - 1,  # №
            item.get('supplier_name', item.get('supplier', 'Unknown')),  # Поставщик
            parser_type,  # Парсер
            item.get('name', ''),  # Наименование
            item.get('article', ''),  # Артикул
            item.get('qty', ''),  # Количество
            item.get('unit', ''),  # Единица
            item.get('price', ''),  # Цена
            item.get('currency', 'RUB'),  # Валюта
            item.get('total', ''),  # Сумма
            f"{item.get('confidence', 0):.1%}",  # Уверенность
            item.get('source', ''),  # Источник
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # Дата обработки
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            
            # Применяем цвет фона
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, 
                                       end_color=fill_color, 
                                       fill_type='solid')
            
            # Форматируем числа
            if col in [6, 8, 10]:  # Количество, Цена, Сумма
                if isinstance(value, (int, float)) and value != '':
                    cell.number_format = '#,##0.00'
            
            # Выравнивание
            if col in [1, 6, 8, 10]:  # Числовые колонки
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
    
    def _get_parser_color(self, parser_type: str) -> Optional[str]:
        """Возвращает цвет для типа парсера"""
        color_map = {
            'supplier_profile': self.colors['supplier_profile'],
            'universal': self.colors['universal'],
            'commercial': self.colors['commercial'],
            'invoice': self.colors['invoice'],
            'competitive': self.colors['competitive']
        }
        return color_map.get(parser_type, None)
    
    def _apply_formatting(self, ws, num_items: int):
        """Применяет форматирование к листу"""
        # Устанавливаем ширину колонок
        column_widths = [5, 20, 15, 40, 15, 12, 10, 15, 8, 15, 12, 20, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Добавляем границы
        thin_border = Border(
            left=Side(style='thin', color=self.colors['border']),
            right=Side(style='thin', color=self.colors['border']),
            top=Side(style='thin', color=self.colors['border']),
            bottom=Side(style='thin', color=self.colors['border'])
        )
        
        # Применяем границы ко всем ячейкам с данными
        for row in range(1, num_items + 2):  # +2 для заголовков
            for col in range(1, 14):  # 13 колонок
                ws.cell(row=row, column=col).border = thin_border
        
        # Замораживаем заголовки
        ws.freeze_panes = 'A2'
    
    def create_backup(self, file_path: str) -> str:
        """Создает резервную копию файла"""
        if not os.path.exists(file_path):
            return None
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = f"{file_path}.backup.{timestamp}"
        
        try:
            shutil.copy2(file_path, backup_path)
            logger.info(f"Создана резервная копия: {backup_path}")
            return backup_path
        except Exception as e:
            logger.error(f"Ошибка создания резервной копии: {e}")
            return None
    
    def write_summary_sheet(self, wb, parser_info: Dict[str, Any], items: List[Dict[str, Any]]):
        """Создает лист с сводкой по обработке"""
        if 'Summary' in wb.sheetnames:
            ws = wb['Summary']
        else:
            ws = wb.create_sheet('Summary')
        
        # Очищаем лист
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
        
        # Записываем сводку
        summary_data = [
            ['Сводка по обработке', ''],
            ['Дата обработки', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['', ''],
            ['Результаты парсинга:', ''],
            ['Лучший парсер', parser_info.get('best_parser', 'Unknown')],
            ['Количество позиций', len(items)],
            ['Общая стоимость', sum(item.get('total', 0) for item in items)],
            ['Средняя уверенность', f"{sum(item.get('confidence', 0) for item in items) / len(items):.1%}" if items else "0%"],
            ['', ''],
            ['Детали по парсерам:', ''],
        ]
        
        # Добавляем информацию о каждом парсере
        parser_results = parser_info.get('parser_results', {})
        for parser_name, result in parser_results.items():
            if isinstance(result, dict) and 'error' not in result:
                summary_data.append([
                    parser_name,
                    f"{result.get('count', 0)} позиций, уверенность {result.get('avg_confidence', 0):.1%}"
                ])
        
        # Записываем данные
        for row_idx, (label, value) in enumerate(summary_data, 1):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)
            
            # Форматируем заголовки
            if row_idx in [1, 4, 10]:
                ws.cell(row=row_idx, column=1).font = Font(bold=True)
                ws.cell(row=row_idx, column=1).fill = PatternFill(
                    start_color=self.colors['header'],
                    end_color=self.colors['header'],
                    fill_type='solid'
                )
        
        # Устанавливаем ширину колонок
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
